from openpyxl import load_workbook
wb = load_workbook('data_preprocessing.xlsx')
print(wb.sheetnames)
ws =  wb.active

rownum = ws.max_row
route_num = 0
old_date = None
all_arcs = []
arcs_in_route = []
for i in range(1, rownum + 1):
    if old_date != ws.cell(row = i, column = 3).value:
        all_arcs.append(arcs_in_route)
        arcs_in_route= []
        arcs_in_route.append(ws.cell(row = i, column = 9).value)
        route_num += 1
        old_date = ws.cell(row = i, column = 3).value
    else:
        arcs_in_route.append(ws.cell(row = i, column = 9).value)
    # display the number of mamber each clustering
print(all_arcs)

all_arcs_updated = []
for rout in all_arcs:
    all_arcs_updated.append(list(set(rout)))

return_list = []

for single_route in all_arcs_updated:
    list_to_add = []
    for ind in range(len(single_route)):
        if ind == 0:
            list_to_add.append((0,single_route[ind]))
        else:
            list_to_add.append((single_route[ind-1], single_route[ind]))
        if ind == len(single_route) -1:
            list_to_add.append((single_route[ind], 0))

    for ele in list_to_add:
        return_list.append(ele)
    
print(return_list)

N = list(range(1, 51))
V = [0] + N
A = [(i,j) for i in V for j in V if i != j]
print(A)
support_data = {(i,j): return_list.count((i,j)) / len(all_arcs)  for i,j in A}

print(support_data)

# Elbow Method for K means
# Import ElbowVisualizer
#from yellowbrick.cluster import KElbowVisualizer
#model = KMeans()
# k is range of number of clusters.
#visualizer = KElbowVisualizer(model, k=(2,30), timings= True)
#visualizer.fit(features)        # Fit data to visualizer
#visualizer.show()        # Finalize and render figure