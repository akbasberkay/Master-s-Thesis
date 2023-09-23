import geopy.distance

import openpyxl

wb_obj = openpyxl.load_workbook('E-commerce original data.xlsx')

sheet_obj = wb_obj.active
starting_row = 3

for i in range(10477):
    latitude_old = sheet_obj.cell(row = starting_row-1, column = 6)
    longitude_old = sheet_obj.cell(row = starting_row-1, column = 7)
    latitude_new = sheet_obj.cell(row = starting_row, column = 6)
    longitude_new = sheet_obj.cell(row = starting_row, column = 7)
    print(latitude_new)
    print(latitude_old)

    coords_1 = (latitude_old, longitude_old)
    coords_2 = (latitude_new, longitude_new)

    print geopy.distance.geodesic(coords_1, coords_2).km
    