import requests
import openpyxl
import xmltodict

wb_obj = openpyxl.load_workbook('Calculate_Distance_and_travel_times.xlsx')

ws = wb_obj.active

berkay = ws.cell(row = 4, column = 1).value
print(berkay)

active_row = 2
x_coordinates = []
y_coordinates = []
while active_row < 53:
    xval = ws.cell(row = active_row, column = 2).value
    yval = ws.cell(row = active_row, column = 3).value
    x_coordinates.append(xval)
    y_coordinates.append(yval)
    active_row += 1
print(x_coordinates)
print(y_coordinates)

num_of_data = 51
N = [i for i in range(51)]
A = [(i, j) for i in N for j in N if i != j]
print(A)
time_dict = {}
distance_dict = {}
for ele in A:
    print(ele)
    print(ele[0])
    print(ele[1])
    url = "https://maps.googleapis.com/maps/api/distancematrix/xml?origins=" + str(x_coordinates[ele[0]]) + "%2C" + str(y_coordinates[ele[0]]) + "&destinations=" + str(x_coordinates[ele[1]]) + "%2C" + str(y_coordinates[ele[1]]) + "&key=AIzaSyAeKf-SA9l8N2YM9aHS_8JqAhH8Dx586Ic"
    print(url)
    payload={}
    headers = {}
    response = requests.request("GET", url, headers=headers, data=payload)

    print(response.text)
    print(type(response.text))

    doc = xmltodict.parse(response.text)
    print(doc)
    print(type(doc))
    time = doc['DistanceMatrixResponse']['row']['element']['duration']['value']
    distance = doc['DistanceMatrixResponse']['row']['element']['distance']['value']
    print(time)
    print(distance)
    time_dict[ele] = time
    distance_dict[ele] = distance

print(time_dict)
print(distance_dict)

""" url = "https://maps.googleapis.com/maps/api/distancematrix/xml?origins=31.17048885%2C121.4334182&destinations=31.1649906%2C121.428579&key=AIzaSyAeKf-SA9l8N2YM9aHS_8JqAhH8Dx586Ic"

payload={}
headers = {}

response = requests.request("GET", url, headers=headers, data=payload)

print(response.text)
print(type(response.text))

doc = xmltodict.parse(response.text)
print(doc)
print(type(doc))
time = doc['DistanceMatrixResponse']['row']['element']['duration']['value']
distance = doc['DistanceMatrixResponse']['row']['element']['distance']['value']
print(time)
print(distance)
 """
""" doc = {'DistanceMatrixResponse': 
    {'status': 'OK', 
     'origin_address': 'China, Shang Hai Shi, Xu Hui Qu, 漕溪路258弄7号 邮政编码: 200235', 
     'destination_address': '91 Xiqin Rd, Xu Hui Qu, Shang Hai Shi, China, 200235', 
     'row': 
        {'element': 
            {'status': 'OK', 
                'duration': 
                {'value': '339', 'text': '6 mins'}, 'distance': {'value': '1210', 'text': '1.2 km'}}}}}

print(doc['DistanceMatrixResponse'])
print(type(doc['DistanceMatrixResponse']))
print(doc['DistanceMatrixResponse']['row']['element']['duration']['value']) """
#['element']['duration']['value']